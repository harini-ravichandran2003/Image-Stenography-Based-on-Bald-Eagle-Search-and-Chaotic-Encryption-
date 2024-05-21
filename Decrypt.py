import numpy as np
import matplotlib.pyplot as plt
from PIL import Image
import pandas as pd
import xlsxwriter
import openpyxl 
import tkinter as tk
from tkinter import filedialog


def chen_scheme(x, y, z):
    x_dot = 35 * (y - x)
    y_dot = -7 * x + x * z + 28 * y
    z_dot = x * y - 3 * z
    return x_dot, y_dot, z_dot

def arraytoexcel(array):
    workbook = xlsxwriter.Workbook('Key.xlsx')
    sheet = workbook.add_worksheet('Key')
    row = np.size(array,axis=0)
    col = np.size(array,axis=1)
    print(row,col)
    for i in range(row):
        for j in range(col):
            sheet.write(i,j,array[i,j])

    workbook.close()

def genkey(image):
    image_size = image.shape

    #print(image_size)

    # Step 1
    P = image
    K0 = np.random.randint(0, 512, size=(4,))
    #print(K0)
    iterations = 50
    k0_1, k0_2, k0_3 = K0[1], K0[2], K0[3]
    
    x = image_size[1]
    y = image_size[0]
    
    # Step 2
    C1, C2, C3 = [], [], []
    for i in range(1001, 1000 + x * y + 1):
        x_dot, y_dot, z_dot = chen_scheme(K0[0], k0_1, k0_2)
        #print(x_dot, y_dot, z_dot)
        K0[0], k0_1, k0_2 = x_dot, y_dot, z_dot
        C1.append(x_dot % 256)
        C2.append(y_dot % 256)
        C3.append(z_dot % 256)

    # Step 3
    #print(image_size[0])
    #print(image_size[1])
   
    A = np.arange(x*y).reshape(x,y)
    A0 = np.zeros_like(A)
    for i in range(iterations):
        for j in range(y):
            if 0 <= C1[i] < 0.25:
                A0 = np.rot90(A, k=1, axes=(0, 1))
            elif 0.25 <= C1[i] < 0.5:
                A0 = np.rot90(A, k=2, axes=(0, 1))
            elif 0.5 <= C1[i] < 0.75:
                A0 = np.rot90(A, k=3, axes=(0, 1))
            else:
                A0 = np.rot90(A, k=4, axes=(0, 1))

    # Step 4
    P0 = np.zeros_like(P)
    for i in range(x):
        for j in range(y):
            index = C2[i * y + j]
            P0[j, i] = P[A0[i, j] % y, A0[i, j] // y]

    P00 = np.zeros_like(P0)
    for i in range(x):
        for j in range(y):
            P00[j, i] = (P0[j, i] + C3[i * y + j]) % 256

    size = P00.shape
    Key = np.arange(x * y).reshape(x,y)
    print(x,y)
    #x = size[0]
    #y = size[1]
    #print(image_size[1],image_size[0])
    for i in range(x):
        for j in range(y):
            Key[i,j] = P00[j,i,0]
    

    #print(Key)
    return Key
def KeyFromExcel(keyfile):
    data = openpyxl.open(keyfile)
    sheet = data['Key']
    row,col = sheet.max_row,sheet.max_column
    key = np.arange(row*col).reshape(row, col)
    m,n=0,0
    for i in range(1,row+1):
        for j in range(1,col+1):
            key[i-1,j-1] = sheet.cell(i,j).value

    return key
    
def decrypt_image(cimage,keyfile):   

    # Import an image from directory: 
    input_image = Image.open(cimage) 
    key = KeyFromExcel(keyfile)
    # Extracting pixel map: 
    #pixel_map = input_image.load() 
    x, y = input_image.size
    #print(row,col)
    dimage = Image.new(mode="RGB", size=(x, y))
    for i in range(x):
        for j in range(y):
            #print(i,j)
            if len(input_image.getpixel((i, j)))==4:
                r,g,b,a = input_image.getpixel((i, j))
            else:
                r,g,b = input_image.getpixel((i, j))
            #print(key[i,j,1])
            #if i==x-1 and j<5:
                #print("Before Dec: ",r,g,b,a,key[i,j])
            r = r ^ key[i,j]
            g = g ^ key[i,j]
            b = b ^ key[i,j]
            
            #if i==x-1 and j<5:
                #print("After Dec: ",r,g,b,a,key[i,j])
            dimage.putpixel((i,j),(r,g,b))
        
    dimage.save('de.png')

    

    
    

# Load an example image (you can replace this with your own image loading code)
#image = plt.imread('lenag_dummy.jpg')

# Encrypt the image

root = tk.Tk()
root.withdraw()

print("Decrypting Image....")
inputfname = filedialog.askopenfilename(title='Select Stego Image')
keyfile = filedialog.askopenfilename(title='Select Key File (Excel)')
print("Input File: "+inputfname)
print("Key File: "+keyfile)
print("Decryption is going on...")
decrypt_image(inputfname, keyfile)
print("Decryption Completed...")
print("Decrypted Image: de.png")




