import cv2
import imageio
import xlsxwriter
from PIL import Image
import openpyxl 
import binascii
import tkinter as tk
from tkinter import filedialog


def ImageToExcel(imfilename, excelfilename):
    workbook = xlsxwriter.Workbook(excelfilename)
    worksheetRed = workbook.add_worksheet('Red')
    worksheetGreen = workbook.add_worksheet('Green')
    worksheetBlue = workbook.add_worksheet('Blue')
    
    # Import an image from directory: 
    input_image = Image.open(imfilename) 
  
    # Extracting pixel map: 
    pixel_map = input_image.load() 
  
    # Extracting the width and height  of the image: 
    width, height = input_image.size 
    
    # taking half of the width: 
    for i in range(width): 
        for j in range(height):
            # getting the RGB pixel value. 
            r, g, b = input_image.getpixel((i, j))  
            worksheetRed.write(i,j,r)
            worksheetGreen.write(i,j,g)
            worksheetBlue.write(i,j,b)
    workbook.close()

def ShowImage(fname):
    img = cv2.imread(fname) 
    cv2.imshow('image', img) 
    cv2.waitKey(0)
    cv2.destroyAllWindows()

def ExcelToImage(excelfilename, imfilename):
    data = openpyxl.open(excelfilename)
    red = data['Red']
    green = data['Green'] 
    blue = data['Blue']

    row,column = red.max_row,red.max_column

    im = Image.new(mode="RGB", size=(row, column))

    for i in range(1,row):
        for j in range(1,column):
            r = red.cell(i,j).value
            g = green.cell(i,j).value
            b = blue.cell(i,j).value       
            im.putpixel((i,j),(r,g,b,255))

    im.save(imfilename)
    print('Image Saved in '+imfilename)

def StringToBin(inputstr):
    # Converting String to binary
    res = ''.join(format(ord(i), '08b') for i in inputstr)
    return res

def LSBEncode(coverimage, secretimage):

    # Import an image from directory: 
    cover_image = Image.open(coverimage) 
    secret_image = Image.open(secretimage)
  
  
    # Extracting the width and height  of the image: 
    width, height = cover_image.size 
    swidth, sheight = secret_image.size
    lenbin = format(swidth,'015b')+format(sheight,'015b')
    print(lenbin)
    bitdepth = 24
    if len(cover_image.getpixel((0, 0)))==4:
        bitdepth = 32
    if bitdepth==24:
        output_image = Image.new(mode="RGB", size=(width, height))
    else:
        output_image = Image.new(mode="RGBA", size=(width, height))
    
    output_image = cover_image.copy()
    
    #code for embedding 30 bits into the image
    i = 0
    j = 0
    for k in range(0,30,6):
        #print("("+str(i)+","+str(j)+"): "+lenbin[k]+""+lenbin[k+1])
        px = cover_image.getpixel((i, j))
        #check for alpha channel is found or not in image.
        if bitdepth==24:
            r, g, b = px
        else:
            r,g,b,a = px
        
        #print("Before:",r, g, b,sep=",")
        
        r = int(format(r>>2,'08b')+str(lenbin[k])+str(lenbin[k+1]),2)
        g = int(format(g>>2,'08b')+str(lenbin[k+2])+str(lenbin[k+3]),2)
        b = int(format(b>>2,'08b')+str(lenbin[k+4])+str(lenbin[k+5]),2)
        
        #print("\tAfter: ",r, g, b,sep=",")
        
        if bitdepth==24:
            output_image.putpixel((i,j),(r,g,b))
        else:
            output_image.putpixel((i,j),(r,g,b,a))   
        j = j+1
        if j==width:
            i = i+1
            j = 0
    print(i)
    print(j)
    #length is the number of bytes to be hide
    #e & f - secret image row col 
    e = 0
    f = 0
    k = 0
    ps = secret_image.getpixel((e, f))
    rs, gs, bs = ps[0],ps[1],ps[2]
    bi = format(rs,'08b')+format(gs,'08b')+format(bs,'08b')
    length = len(bi)
    f = f+1
    if f==sheight:
        f=0
        e = e+1
    while i<width:
        while j<height:
            pc = cover_image.getpixel((i, j))
            if bitdepth==24:
                r, g, b = pc
            else:
                r,g,b,a = pc
            if k<24:
                r = int(format(r>>2,'08b')+str(bi[k])+str(bi[k+1]),2)
                k = k + 2
            if k==24:
                if e<swidth:
                    ps = secret_image.getpixel((e, f))
                    rs, gs, bs = ps[0],ps[1],ps[2]
                    bi = format(rs,'08b')+format(gs,'08b')+format(bs,'08b')
                    length = len(bi)
                    k = 0
                    f = f+1
                    if f==sheight:
                        f=0
                        e = e+1
            if k<24:       
                g = int(format(g>>2,'08b')+str(bi[k])+str(bi[k+1]),2)
                k = k + 2
            if k==24:
                if e<swidth:
                    ps = secret_image.getpixel((e, f))
                    rs, gs, bs = ps[0],ps[1],ps[2]
                    bi = format(rs,'08b')+format(gs,'08b')+format(bs,'08b')
                    length = len(bi)
                    k = 0
                    f = f+1
                    if f==sheight:
                        f = 0
                        e = e+1
            if k<24:       
                b = int(format(b>>2,'08b')+str(bi[k])+str(bi[k+1]),2)
                k = k + 2
            if k==24:
                
                if e<swidth:
                    ps = secret_image.getpixel((e, f))
                    rs, gs, bs = ps[0],ps[1],ps[2]
                    bi = format(rs,'08b')+format(gs,'08b')+format(bs,'08b')
                    length = len(bi)
                    k = 0
                    f = f+1
                    if f==sheight:
                        f = 0
                        e = e+1
            if bitdepth==24:
                px = (r,g,b)
            else:
                px = (r,g,b,a)
            output_image.putpixel((i,j),px)
            if e==swidth:
                j = height
                i = width
            j = j + 1
        j = 0
        i = i + 1
    
 

    output_image.save('steg_img.png')  
    print("\nEmbedded Successfully...")
    print("Stego-Image: Saved as steg_img.png...")

def LSBDecode(stegoimage):
    
    input_image = Image.open(stegoimage)
    width, height = input_image.size 
    bitdepth = 24
    if len(input_image.getpixel((0, 0)))==4:
        bitdepth = 32
   
    i = 0
    j = 0
    s=''
    for k in range(0,30,6):
        #print("("+str(i)+","+str(j)+"): "+lenbin[k]+""+lenbin[k+1])
        if bitdepth==24:
            r, g, b = input_image.getpixel((i, j))
        else:
            r, g, b, a = input_image.getpixel((i, j))
            
        #print(":Before:",r, g, b,sep=",")
        r1 = format(r,'08b')
        g1 = format(g,'08b')
        b1 = format(b,'08b')
        s = s+r1[6]+r1[7]+g1[6]+g1[7]+b1[6]+b1[7]      

        #print(":After: ",r1, g1, b1,s,sep=",")
        j = j+1
        if j==width:
            i = i+1
            j = 0
    swidth = int(s[:15],2) #first 15bits
    sheight = int(s[15:],2)  #last 15bits
    output_image = Image.new(mode="RGB", size=(swidth, sheight))
    print(swidth)
    print(sheight)
    s = ''
    e = 0
    f = 0
    while i<width:
        while j<height:
            if e==swidth:
                break            
            if bitdepth==24:
                r, g, b = input_image.getpixel((i, j))
            else:
                r, g, b, a = input_image.getpixel((i, j))
            r1 = format(r,'08b')
            g1 = format(g,'08b')
            b1 = format(b,'08b')
            s = s+r1[6]+r1[7]
            if len(s)==24:
                rs = int(s[0:8],2)
                gs = int(s[8:16],2)
                bs = int(s[16:24],2)
                output_image.putpixel((e,f),(rs,gs,bs))
                f = f + 1
                if f==sheight:
                    f = 0
                    e = e + 1
                s=''
            s = s+g1[6]+g1[7]
            if len(s)==24:
                rs = int(s[0:8],2)
                gs = int(s[8:16],2)
                bs = int(s[16:24],2)
                output_image.putpixel((e,f),(rs,gs,bs))
                f = f + 1
                if f==sheight:
                    f = 0
                    e = e + 1
                s=''
            s = s+b1[6]+b1[7]
            if len(s)==24:
                rs = int(s[0:8],2)
                gs = int(s[8:16],2)
                bs = int(s[16:24],2)
                output_image.putpixel((e,f),(rs,gs,bs))
                f = f + 1
                if f==sheight:
                    f = 0
                    e = e + 1
                s=''
            j = j + 1
        j = 0
        i = i + 1
    output_image.save('en1.png')
    print("\nDecoded Successfully...")
    print("\nSaved As de.png...")

coverimage = filedialog.askopenfilename(title='Select Cover Image')
secretimage =  filedialog.askopenfilename(title='Select Secret Image (Encrypted)')   
LSBEncode(coverimage,secretimage)
